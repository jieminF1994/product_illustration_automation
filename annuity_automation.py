"""
annuity_automation.py
=====================
Automates Corebridge Index Annuity Hypothetical Illustration Tool.

Phases
------
  1  Populate .xlsx workbooks from product_structure.json
  2  Gather Report-tab outputs into tool_calc_output.json
  3  Compare tool_calc_output.json vs product_structure.json → check_report.csv

Recalculation (between Phase 1 and 2) is handled by recalc_helper.py
which uses xlwings to open/recalc/save each workbook for all 3 scenarios.

Usage
-----
  # Run all phases in one submission. By default the template/json are read
  # from the parent folder and results are written under ../results/YYYY-MM-DD/.
  python annuity_automation.py --phases 1 2 3

  # Run only selected PDF test cases
  python annuity_automation.py --phases 1 2 3 \
      --test-cases "Example 1.pdf" "Example 2.pdf"

  # Run from an IDE by setting values in config = { ... } and executing the file
"""

from __future__ import annotations

import argparse
import copy
import csv
import datetime as dt
import json
import logging
import os
import re
import shutil
import string
import subprocess
import sys
import time
import uuid
import warnings
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

from extract_annuity_data import AnnuityExtractorPipeline

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
LOG_FORMAT = "%(asctime)s [%(levelname)s] [%(name)s] %(message)s"
log = logging.getLogger("annuity_automation")


def configure_logging(log_file: Path | None = None):
    root = logging.getLogger()
    root.setLevel(logging.INFO)
    formatter = logging.Formatter(LOG_FORMAT)

    has_console = any(getattr(handler, "_annuity_console", False) for handler in root.handlers)
    if not has_console:
        console_handler = logging.StreamHandler(sys.stdout)
        console_handler.setFormatter(formatter)
        console_handler._annuity_console = True  # type: ignore[attr-defined]
        root.addHandler(console_handler)

    if log_file is None:
        return

    resolved_log_file = log_file.resolve()
    has_file = any(
        isinstance(handler, logging.FileHandler)
        and Path(getattr(handler, "baseFilename", "")).resolve() == resolved_log_file
        for handler in root.handlers
    )
    if not has_file:
        resolved_log_file.parent.mkdir(parents=True, exist_ok=True)
        file_handler = logging.FileHandler(resolved_log_file, encoding="utf-8")
        file_handler.setFormatter(formatter)
        root.addHandler(file_handler)

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_PARENT_DIR = SCRIPT_DIR.parent
DEFAULT_TEMPLATE_PATH = PROJECT_PARENT_DIR / "Index Annuity Hypo Illustrations Tool_v1.47 try.xlsx"
DEFAULT_JSON_PATH = PROJECT_PARENT_DIR / "product_structure.json"
DEFAULT_RESULTS_ROOT = PROJECT_PARENT_DIR / "results"
DEFAULT_PHASES = [1, 2, 3]
VALID_PHASES = {1, 2, 3}

# IDE entrypoint configuration.
# Fill values here and run the script directly from your IDE when desired.
config: dict[str, Any] = {
    # "template": str(DEFAULT_TEMPLATE_PATH),
    # "json": str(DEFAULT_JSON_PATH),
    # "pdf_dir": "/absolute/path/to/pdf_dir",
    # "pdf_extractor": "auto",
    # "phases": [1, 2, 3],
    # "test_cases": ["Example 1.pdf", "Example 2.pdf"],
    # "results_root": str(DEFAULT_RESULTS_ROOT),
    # "output_dir": "/absolute/path/to/existing/run_folder",
}

# ---------------------------------------------------------------------------
# openpyxl compatibility (Strict OOXML .xlsx -> Transitional OOXML namespaces)
# ---------------------------------------------------------------------------

STRICT_TO_TRANSITIONAL_URIS: tuple[tuple[bytes, bytes], ...] = (
    (
        b"http://purl.oclc.org/ooxml/spreadsheetml/main",
        b"http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    ),
    (
        b"http://purl.oclc.org/ooxml/officeDocument/relationships",
        b"http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    ),
    (
        b"http://purl.oclc.org/ooxml/drawingml/main",
        b"http://schemas.openxmlformats.org/drawingml/2006/main",
    ),
    (
        b"http://purl.oclc.org/ooxml/drawingml/chart",
        b"http://schemas.openxmlformats.org/drawingml/2006/chart",
    ),
)


def _normalize_strict_xlsx_inplace(xlsx_path: Path) -> bool:
    """
    Rewrite strict OOXML namespace URIs in XML/.rels parts so openpyxl can read
    workbooks exported in strict mode. Returns True if any part changed.
    """
    tmp_path = xlsx_path.with_suffix(xlsx_path.suffix + ".tmp")
    changed = False

    with zipfile.ZipFile(xlsx_path, "r") as zin, zipfile.ZipFile(tmp_path, "w", compression=zipfile.ZIP_DEFLATED) as zout:
        for info in zin.infolist():
            data = zin.read(info.filename)
            if info.filename.endswith(".xml") or info.filename.endswith(".rels"):
                new_data = data
                for old_uri, new_uri in STRICT_TO_TRANSITIONAL_URIS:
                    new_data = new_data.replace(old_uri, new_uri)
                if new_data != data:
                    changed = True
                    data = new_data
            zout.writestr(info, data)

    if changed:
        tmp_path.replace(xlsx_path)
    else:
        try:
            tmp_path.unlink()
        except FileNotFoundError:
            pass
    return changed


def load_workbook_compat(path: Path, **kwargs):
    """
    Load workbook with openpyxl; if strict OOXML leads to zero visible sheets,
    normalize namespaces in-place and retry once.
    """
    def _safe_load():
        with warnings.catch_warnings():
            warnings.filterwarnings(
                "ignore",
                message=r"Cell .* is marked as a date but the serial value .* is outside the limits for dates\.",
                category=UserWarning,
                module=r"openpyxl\.worksheet\._reader",
            )
            return load_workbook(path, **kwargs)

    wb = _safe_load()
    if wb.sheetnames:
        return wb
    wb.close()

    if _normalize_strict_xlsx_inplace(path):
        log.info("Normalized strict OOXML namespaces for %s", path.name)
        wb = _safe_load()
        if wb.sheetnames:
            return wb
        wb.close()

    raise RuntimeError(
        f"Workbook could not be read by openpyxl (no worksheets): {path}"
    )

# ---------------------------------------------------------------------------
# Mapping tables
# ---------------------------------------------------------------------------

# JSON product name  →  Excel product name  (exact string in Product named range)
PRODUCT_MAP: dict[str, str] = {
    # Advisory Series
    "Power Index Advisory":                          "Power Index Advisory",
    "Power Index Elite Advisory":                    "Power Index Elite Advisory",
    "Power Select Advisory":                         "Power Select Advisory",
    # Index Series
    "AG Choice Index 10":                            "AG Choice Index 10",
    "AG Choice Index 10 Plus Income":                "AG Choice Index 10 Plus Income",
    "Power Index 5 Plus":                            "Power Index 5 Plus (Standard)",
    "Power Index 5 Plus (Standard)":                 "Power Index 5 Plus (Standard)",
    "Power Index 5 Plus (Wells Fargo)":              "Power Index 5 Plus (Wells Fargo)",
    "Power Index 5 Elite":                           "Power Index 5 Elite",
    "Power Index 5 Elite Plus Income":               "Power Index 5 Elite Plus Income",
    "Power Index 7 Plus":                            "Power Index 7 Plus (Wells Fargo)",
    "Power Index 7 Plus Income":                     "Power Index 7 Plus Income (Wells Fargo)",
    "Power Index Elite":                             "Power Index Elite",
    "Power Index Elite Plus Income":                 "Power Index Elite Plus Income",
    "Power Index Plus":                              "Power Index Plus",
    "Power Index Plus Income":                       "Power Index Plus Income",
    "Power Index Preferred":                         "Power Index Preferred",
    "Power Index Preferred Plus Income":             "Power Index Preferred Plus Income",
    "Power Index Advisory Income":                   "Power Index Advisory Income",
    "Power Index Premier":                           "Power Index Premier",
    "Power Index Premier Plus Income":               "Power Index Premier Plus Income",
    # Power Index 5 NY
    "Power Index 5 NY":                              "Power Index 5 NY",
    "Power Index 5 NY with LIB":                     "Power Index 5 NY with LIB",
    "Power Index Premier NY":                        "Power Index Premier NY",
    "Power Index Premier NY with LIB":               "Power Index Premier NY with LIB",
    # Power Select Series
    "Power Select Builder":                          "Power Select Builder",
    "Power Select Builder 8":                        "Power Select Builder 8",
    "Power Select Plus Income":                      "Power Select Plus Income",
    # Shelf / Power Protector
    "Power 10 Protector":                            "Power 10 Protector",
    "Power 10 Protector Plus Income":                "Power 10 Protector Plus Income - Max Income",
    "Power 10 Protector Plus Income - Level Income": "Power 10 Protector Plus Income - Level Income",
    "Power 10 Protector Plus Income - Max Income":   "Power 10 Protector Plus Income - Max Income",
    "Power 5 Protector":                             "Power 5 Protector",
    "Power 7 Protector":                             "Power 7 Protector",
    "Power 7 Protector Plus Income":                 "Power 7 Protector Plus Income - Max Income",
    "Power 7 Protector Plus Income - Level Income":  "Power 7 Protector Plus Income - Level Income",
    "Power 7 Protector Plus Income - Max Income":    "Power 7 Protector Plus Income - Max Income",
    # Shelf Index Series
    "Power Advantage 10":                            "Power Advantage 10",
    "Power Advantage 10 Plus Income":                "Power Advantage 10 Plus Income",
    "Power Advantage 7":                             "Power Advantage 7",
    "Power Advantage 7 Plus Income":                 "Power Advantage 7 Plus Income",
    # AICO
    "Power Select AICO":                             "Power Select AICO",
}

# Excel product name  →  Category (AG column in Ref tab)
PRODUCT_TO_CATEGORY: dict[str, str] = {
    "Power Index Advisory":                          "Advisory Series",
    "Power Index Elite Advisory":                    "Advisory Series",
    "Power Select Advisory":                         "Advisory Series",
    "AG Choice Index 10":                            "Index Series",
    "AG Choice Index 10 Plus Income":                "Index Series",
    "Power Index 5 Plus (Standard)":                 "Index Series",
    "Power Index 5 Plus (Wells Fargo)":              "Index Series",
    "Power Index 5 Plus (Chase)":                    "Index Series",
    "Power Index 5 Plus Income (Wells Fargo)":       "Index Series",
    "Power Index 5 Elite":                           "Index Series",
    "Power Index 5 Elite Plus Income":               "Index Series",
    "Power Index 7 Plus (Wells Fargo)":              "Index Series",
    "Power Index 7 Plus Income (Wells Fargo)":       "Index Series",
    "Power Index Elite":                             "Index Series",
    "Power Index Elite Plus Income":                 "Index Series",
    "Power Index Plus":                              "Index Series",
    "Power Index Plus Income":                       "Index Series",
    "Power Index Preferred":                         "Index Series",
    "Power Index Preferred Plus Income":             "Index Series",
    "Power Index Advisory Income":                   "Index Series",
    "Power Index Elite Advisory with GLB":           "Index Series",
    "Power Index Premier":                           "Index Series",
    "Power Index Premier Plus Income":               "Index Series",
    "Power Index 5 NY":                              "Power Index 5 NY",
    "Power Index 5 NY with LIB":                     "Power Index 5 NY",
    "Power Index Premier NY":                        "Power Index Premier NY",
    "Power Index Premier NY with LIB":               "Power Index Premier NY",
    "Power Select Builder":                          "Power Select Series",
    "Power Select Builder 8":                        "Power Select Series",
    "Power Select Plus Income":                      "Power Select Series",
    "Power Select Advisory with GLB":                "Power Select Series",
    "Power 10 Protector":                            "Shelf & Power Protector Series",
    "Power 10 Protector Plus Income - Level Income": "Shelf & Power Protector Series",
    "Power 10 Protector Plus Income - Max Income":   "Shelf & Power Protector Series",
    "Power 5 Protector":                             "Shelf & Power Protector Series",
    "Power 7 Protector":                             "Shelf & Power Protector Series",
    "Power 7 Protector Plus Income - Level Income":  "Shelf & Power Protector Series",
    "Power 7 Protector Plus Income - Max Income":    "Shelf & Power Protector Series",
    "Power Advantage 10":                            "Shelf Index Series",
    "Power Advantage 10 Plus Income":                "Shelf Index Series",
    "Power Advantage 7":                             "Shelf Index Series",
    "Power Advantage 7 Plus Income":                 "Shelf Index Series",
    "Power Select AICO":                             "Power Select AICO",
}

# JSON strategy name  →  Excel strategy name (col E in the rate table, rows 142+)
STRATEGY_MAP: dict[str, str] = {
    # Fixed
    "Fixed Account":                                 "1-Year Fixed Rate, Fixed Account",
    "1-Year Fixed Rate":                             "1-Year Fixed Rate, Fixed Account",
    "1-Year Fixed Rate, Fixed Account":              "1-Year Fixed Rate, Fixed Account",
    # S&P 500 cap
    "S&P Annual PTP with Cap":                       "1-Year PTP with Cap, S&P 500",
    "1-Year PTP with Cap, S&P 500":                  "1-Year PTP with Cap, S&P 500",
    "S&P 500 Annual PTP Cap":                        "1-Year PTP with Cap, S&P 500",
    # S&P 500 secure cap
    "S&P Annual PTP with Secure Cap":                "1-Year PTP with Secure Cap, S&P 500",
    "1-Year PTP with Secure Cap, S&P 500":           "1-Year PTP with Secure Cap, S&P 500",
    # S&P 500 par rate
    "S&P Annual PTP with ParRate":                   "1-Year PTP with ParRate, S&P 500",
    "1-Year PTP with ParRate, S&P 500":              "1-Year PTP with ParRate, S&P 500",
    "S&P Annual PTP Participation Rate":             "1-Year PTP with ParRate, S&P 500",
    # PIMCO
    "PIMCO Annual PTP with Cap":                     "1-Year PTP with Cap, PIMCO Global Optima",
    "1-Year PTP with Cap, PIMCO Global Optima":      "1-Year PTP with Cap, PIMCO Global Optima",
    "PIMCO Global Optima PTP Cap":                   "1-Year PTP with Cap, PIMCO Global Optima",
    "PIMCO Annual PTP with ParRate":                 "1-Year PTP with ParRate, PIMCO Global Optima",
    "1-Year PTP with ParRate, PIMCO Global Optima":  "1-Year PTP with ParRate, PIMCO Global Optima",
    "PIMCO Annual PTP Participation Rate":           "1-Year PTP with ParRate, PIMCO Global Optima",
    # MLSB
    "MLSB Annual PTP with ParRate":                  "1-Year PTP with ParRate, MLSB",
    "1-Year PTP with ParRate, MLSB":                 "1-Year PTP with ParRate, MLSB",
    "MLSB Annual PTP with Cap":                      "1-Year PTP with Cap, MLSB",
    "MLSB Annual PTP Participation Rate":            "1-Year PTP with ParRate, MLSB",
    # Franklin
    "Franklin Annual PTP with ParRate":              "1-Year PTP with ParRate, Franklin",
    "1-Year PTP with ParRate, Franklin":             "1-Year PTP with ParRate, Franklin",
    # Russell
    "Russell 2000 Annual PTP with Cap":              "1-Year PTP with Cap, Russell 2000",
    "1-Year PTP with Cap, Russell 2000":             "1-Year PTP with Cap, Russell 2000",
    "Russell 2000 Annual PTP with Rate Cap":         "1-Year PTP with Cap, Russell 2000",
    # MSCI
    "MSCI Annual PTP with Cap":                      "1-Year PTP with Cap, MSCI",
    "1-Year PTP with Cap, MSCI":                     "1-Year PTP with Cap, MSCI",
    # 5-year
    "5-Year PTP with Cap, S&P 500":                  "5-Year PTP with Cap, S&P 500",
    "S&P 5-Year PTP with Cap":                       "5-Year PTP with Cap, S&P 500",
    # Performance-triggered
    "1-Year Performance-Triggered Account, S&P 500": "1-Year Performance-Triggered Account, S&P 500",
    "S&P Performance Triggered":                     "1-Year Performance-Triggered Account, S&P 500",
    "S&P Annual PTP Performance-Triggered":          "1-Year Performance-Triggered Account, S&P 500",
    "Triggered":                                     "1-Year Performance-Triggered Account, S&P 500",
}

# JSON Living Benefit  →  Excel LivBen value
LIVBEN_MAP: dict[str, str] = {
    "none":                              "None",
    "lifetime income choice":            "Lifetime Income Choice",
    "lifetime income max":               "Lifetime Income Max",
    "lifetime income plus flex":         "Lifetime Income Plus Flex",
    "lifetime income plus multiplier flex": "Lifetime Income Plus Multiplier Flex",
    "lifetime income plus":              "Lifetime Income Plus Flex",
    "lifetime income plus multiplier":   "Lifetime Income Plus Multiplier Flex",
    "lifetime income multiplier flex":   "Lifetime Income Plus Multiplier Flex",
    "income max":                        "Lifetime Income Max",
    "income plus flex":                  "Lifetime Income Plus Flex",
}

# Rate type detection per strategy name
def _detect_rate_column(strategy_excel: str) -> str:
    """Return which rate column ('cap','parrate','spread','triggered','fixed') this strategy uses."""
    s = strategy_excel.lower()
    if "fixed rate" in s:
        return "fixed"
    if "parrate" in s:
        return "parrate"
    if "spread" in s:
        return "spread"
    if "triggered" in s or "performance-triggered" in s:
        return "triggered"
    if "cap" in s:
        return "cap"
    return "cap"

# Excel rate-table column letters for Inputs & Summary rows 141+
RATE_COL_MAP = {
    "fixed":     "F",   # Fixed Rate column
    "parrate":   "G",   # ParRate column  (note: row 141 header says F=ParRate, G=Cap)
    "cap":       "H",   # Cap column       BUT actual layout: E=strategy name, F=Fixed Rate col header
    "spread":    "I",   # Spread
    "triggered": "J",   # Triggered Rate
}
# Actual layout from inspection: Row 141 headers are E=Fixed Rate, F=ParRate, G=Cap, H=Spread, I=Triggered Rate
# Allocation is in col C (same rows)
RATE_COL_ACTUAL = {
    "fixed":     "E",
    "parrate":   "F",
    "cap":       "G",
    "spread":    "H",
    "triggered": "I",
}
ALLOC_COL = "C"  # allocation percentage column in strategy rate rows

# Contract type mapping from JSON → Excel
CONTRACT_TYPE_MAP = {
    "qualified":     "Qualified",
    "non-qualified": "Non-Qualified",
    "nonqualified":  "Non-Qualified",
    "non qualified": "Non-Qualified",
    "ira":           "Qualified",
    "roth":          "Non-Qualified",
}

ELECTION_MAP = {
    "single":       "Single Life",
    "single life":  "Single Life",
    "joint":        "Joint Life",
    "joint life":   "Joint Life",
    "none":         "None",
}

STATE_ABBREV = {
    "alabama": "AL","alaska": "AK","arizona": "AZ","arkansas": "AR","california": "CA",
    "colorado": "CO","connecticut": "CT","delaware": "DE","florida": "FL","georgia": "GA",
    "hawaii": "HI","idaho": "ID","illinois": "IL","indiana": "IN","iowa": "IA",
    "kansas": "KS","kentucky": "KY","louisiana": "LA","maine": "ME","maryland": "MD",
    "massachusetts": "MA","michigan": "MI","minnesota": "MN","mississippi": "MS",
    "missouri": "MO","montana": "MT","nebraska": "NE","nevada": "NV",
    "new hampshire": "NH","new jersey": "NJ","new mexico": "NM","new york": "NY",
    "north carolina": "NC","north dakota": "ND","ohio": "OH","oklahoma": "OK",
    "oregon": "OR","pennsylvania": "PA","rhode island": "RI","south carolina": "SC",
    "south dakota": "SD","tennessee": "TN","texas": "TX","utah": "UT",
    "vermont": "VT","virginia": "VA","washington": "WA","west virginia": "WV",
    "wisconsin": "WI","wyoming": "WY",
}

# ---------------------------------------------------------------------------
# Dataclasses
# ---------------------------------------------------------------------------

@dataclass
class StrategyInput:
    excel_name: str       # exact name as it appears in the rate table col E
    allocation: float     # 0-1
    rate: float           # e.g. 0.09 for 9%
    rate_type: str        # cap / parrate / spread / triggered / fixed


@dataclass
class TestCaseInputs:
    pdf_name: str
    # named-range values
    category: str = ""
    product: str = ""
    livben: str = "None"
    summary: str = "Specific"
    ssubcat: str = "Standard"
    mcirind: str = "No"
    iss_age1: int = 65
    iss_age2: int | None = None
    mat_age: int = 95
    state: str = ""
    contract_type: str = "Qualified"
    election: str = "Single Life"
    rmd_ind: str = "No"
    cl_prem: float = 10000.0
    pfed: str = "12/31/2025"       # performance focus end date
    mwv_rate: float = 0.0
    cgr: float = 0.03
    gmir: float = 0.0025
    gmabc: str = "CDSC"
    gmabr: float = 0.05
    wd_rate: float = 1.0
    lia_dur: int = 1
    sw_ind: str = "No"
    sw_mode: str = "Fixed Amount"
    sw_amount: float = 500.0
    sw_rate: float = 0.05
    sw_start: int = 0
    sw_end: int = 29
    aico_pr_fee: float = 0.008
    omr: float = 2.0
    omax_r: float = 1.0
    icp: int = 100
    girm1: float = 2.0
    girm2: float = 1.0
    nirm1: float = 1.0
    nirm2: float = 1.0
    adv_fee: float = 0.0
    max_igp: int = 10
    strategies: list[StrategyInput] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)

# ---------------------------------------------------------------------------
# Phase 1: TestCaseParser
# ---------------------------------------------------------------------------

class TestCaseParser:
    """Parse one JSON test case dict → TestCaseInputs, logging errors."""

    def __init__(self, pdf_name: str, tc: dict):
        self.pdf_name = pdf_name
        self.tc = tc
        self.errors: list[str] = []

    def _get(self, path: list[str], default=None):
        """Navigate nested dict safely (case/format insensitive key matching)."""
        def _norm_key(k: Any) -> str:
            return re.sub(r"[^a-z0-9]", "", str(k).lower())

        node = self.tc
        for key in path:
            if not isinstance(node, dict):
                return default
            if key in node:
                node = node[key]
                continue
            want = _norm_key(key)
            matched = None
            for k in node.keys():
                if _norm_key(k) == want:
                    matched = k
                    break
            if matched is None:
                return default
            node = node[matched]
        return node

    def _warn_missing(self, name: str):
        msg = f"{name} missing from PDF, please check PDF and verify if it so"
        self.errors.append(msg)
        log.warning("[%s] %s", self.pdf_name, msg)

    def _warn_invalid(self, name: str):
        msg = f"{name}'s value is invalid from PDF, please check PDF and verify if it so"
        self.errors.append(msg)
        log.warning("[%s] %s", self.pdf_name, msg)

    def _has_living_benefit(self, raw_lb: Any) -> bool:
        val = str(raw_lb or "").strip().lower()
        return bool(val) and "none" not in val

    def _product_with_income_variant(self, raw_product: str, raw_lb: Any) -> str:
        if not raw_product or not self._has_living_benefit(raw_lb):
            return raw_product

        product_lower = raw_product.lower()
        if any(token in product_lower for token in ("plus income", "advisory income", "with lib", "max income", "level income")):
            return raw_product

        candidates = [
            f"{raw_product} Plus Income",
            f"{raw_product} with LIB",
        ]
        if re.search(r"(?i)\bplus\b", raw_product) and "plus income" not in product_lower:
            candidates.append(re.sub(r"(?i)\bplus\b", "Plus Income", raw_product, count=1))
        if "advisory" in raw_product.lower() and "income" not in raw_product.lower():
            candidates.append(re.sub(r"advisory", "Advisory Income", raw_product, flags=re.IGNORECASE))

        for candidate in candidates:
            if candidate in PRODUCT_MAP:
                log.info(
                    "[%s] Adjusting product from '%s' to '%s' because living benefit is present",
                    self.pdf_name,
                    raw_product,
                    candidate,
                )
                return candidate

        return raw_product

    def _strategy_records(self, ics: Any) -> list[tuple[str, dict[str, Any]]]:
        """
        Normalize interest_crediting_strategy payloads into
        [(strategy_name, strategy_details_dict), ...].

        Supported shapes:
        - {"Strategy A": {"Allocation": "...", "Cap": "..."}, ...}
        - {
              "Strategy": ["Strategy A", "Strategy B"],
              "Allocation": ["50%", "50%"],
              "Rate": ["5.00%", "N/A"],
              ...
          }
        - [{"Strategy": "Strategy A", ...}, ...]
        """
        if not ics:
            return []

        if isinstance(ics, list):
            records: list[tuple[str, dict[str, Any]]] = []
            for item in ics:
                if not isinstance(item, dict):
                    continue
                strategy_name = (
                    item.get("Strategy")
                    or item.get("strategy")
                    or item.get("Name")
                    or item.get("name")
                )
                if strategy_name:
                    records.append((str(strategy_name), item))
            return records

        if not isinstance(ics, dict):
            return []

        if any(isinstance(v, dict) for v in ics.values()):
            return [(str(k), v) for k, v in ics.items() if isinstance(v, dict)]

        strategy_names = (
            ics.get("Strategy")
            or ics.get("strategy")
            or ics.get("Interest_Crediting_Strategy")
            or ics.get("interest_crediting_strategy")
        )
        if strategy_names is None:
            return []

        if not isinstance(strategy_names, list):
            strategy_names = [strategy_names]

        records = []
        for idx, strategy_name in enumerate(strategy_names):
            if strategy_name in (None, ""):
                continue
            details: dict[str, Any] = {}
            for key, value in ics.items():
                if re.sub(r"[^a-z0-9]", "", str(key).lower()) == "strategy":
                    continue
                if isinstance(value, list):
                    details[key] = value[idx] if idx < len(value) else None
                else:
                    details[key] = value
            records.append((str(strategy_name), details))
        return records

    def parse(self) -> TestCaseInputs:
        inp = TestCaseInputs(pdf_name=self.pdf_name)
        p = self._get(["profile"]) or self._get(["Profile"]) or {}
        inc = self._get(["income_details"]) or {}
        ics = self._get(["interest_crediting_strategy"]) or {}
        raw_lb = inc.get("Living_Benefit") or inc.get("living_benefit") or "None"

        # --- Product & Category ---
        raw_product = p.get("Product") or p.get("product") or ""
        if not raw_product:
            self._warn_missing("Product")
        raw_product = self._product_with_income_variant(raw_product, raw_lb)
        excel_product = PRODUCT_MAP.get(raw_product, "")
        if raw_product and not excel_product:
            # fuzzy fallback: case-insensitive contains match
            for k, v in PRODUCT_MAP.items():
                if raw_product.lower() in k.lower() or k.lower() in raw_product.lower():
                    excel_product = v
                    break
        if not excel_product:
            self._warn_invalid("Product")
            excel_product = raw_product  # last resort
        inp.product = excel_product
        inp.category = PRODUCT_TO_CATEGORY.get(excel_product, "Index Series")

        # --- Living Benefit ---
        raw_lb_lower = str(raw_lb).lower().strip()
        if "none" in raw_lb_lower or raw_lb_lower == "":
            inp.livben = "None"
        else:
            mapped = LIVBEN_MAP.get(raw_lb_lower)
            if not mapped:
                # fuzzy
                for k, v in LIVBEN_MAP.items():
                    if k in raw_lb_lower or raw_lb_lower in k:
                        mapped = v
                        break
            if mapped:
                inp.livben = mapped
            else:
                self._warn_invalid("LivBen")
                inp.livben = "None"

        # --- Issue Age ---
        raw_age = p.get("Issue_Age") or p.get("issue_age") or p.get("Age")
        if raw_age is None:
            self._warn_missing("IssAge1")
        else:
            try:
                inp.iss_age1 = int(raw_age)
                inp.iss_age2 = int(raw_age)
            except (ValueError, TypeError):
                self._warn_invalid("IssAge1")

        # --- MatAge: IssAge + projection_length ---
        proj_len = p.get("Projection_Length") or p.get("projection_length") or 30
        try:
            inp.mat_age = inp.iss_age1 + int(proj_len)
        except Exception:
            inp.mat_age = inp.iss_age1 + 30

        # --- State ---
        raw_state = p.get("State") or p.get("state") or ""
        if not raw_state:
            self._warn_missing("State")
        else:
            state_upper = raw_state.strip().upper()
            if len(state_upper) == 2:
                inp.state = state_upper
            else:
                abbrev = STATE_ABBREV.get(raw_state.strip().lower(), "")
                if abbrev:
                    inp.state = abbrev
                else:
                    self._warn_invalid("State")
                    inp.state = raw_state[:2].upper()

        # --- Contract Type ---
        raw_ct = p.get("Contract_Type") or p.get("contract_type") or "Qualified"
        inp.contract_type = CONTRACT_TYPE_MAP.get(raw_ct.lower().strip(), "Qualified")

        # --- Election ---
        raw_el = p.get("Election") or p.get("election") or "Single Life"
        inp.election = ELECTION_MAP.get(raw_el.lower().strip(), "Single Life")

        # --- Premium ---
        raw_prem = p.get("Premium") or p.get("premium")
        if raw_prem is None:
            self._warn_missing("CLPrem")
        else:
            try:
                inp.cl_prem = float(str(raw_prem).replace(",", "").replace("$", ""))
            except ValueError:
                self._warn_invalid("CLPrem")

        # --- Withdrawal Rate (MWVRate) ---
        raw_wr = inc.get("Withdrawal_Rate") or inc.get("withdrawal_rate")
        if raw_wr is not None:
            try:
                v = float(str(raw_wr).replace("%", ""))
                inp.mwv_rate = v / 100 if v > 1 else v
            except ValueError:
                self._warn_invalid("MWVRate")

        # --- PFED (index_end_date) ---
        raw_pfed = (self._get(["scenario", "specific", "index_end_date"])
                    or self._get(["scenario", "index_end_date"]))
        inp.pfed = raw_pfed if raw_pfed else "12/31/2025"

        # --- CGR (constant_growth_rate) ---
        raw_cgr = (self._get(["scenario", "constant_growth", "constant_growth_rate"])
                   or self._get(["scenario", "constant_growth_rate"]))
        if raw_cgr is not None:
            try:
                v = float(str(raw_cgr).replace("%", ""))
                inp.cgr = v / 100 if v > 1 else v
            except ValueError:
                self._warn_invalid("CGR")
        else:
            inp.cgr = 0.03

        # --- AICO fields ---
        raw_aico = inc.get("AICO_Fee")
        if raw_aico is not None:
            try:
                v = float(str(raw_aico).replace("%", ""))
                inp.aico_pr_fee = v / 100 if v > 1 else v
            except ValueError:
                self._warn_invalid("AICOPrFee")

        raw_omr = inc.get("AICO_Multiplier_Rate")
        if raw_omr is not None:
            try:
                v = float(str(raw_omr).replace("%", ""))
                inp.omr = v / 100 if v > 1 else v
            except ValueError:
                self._warn_invalid("OMR")

        raw_omax = inc.get("AICO_Maximum_Rate")
        if raw_omax is not None:
            try:
                v = float(str(raw_omax).replace("%", ""))
                inp.omax_r = v / 100 if v > 1 else v
            except ValueError:
                self._warn_invalid("OMaxR")

        # --- Strategies ---
        inp.strategies = self._parse_strategies(ics)
        inp.errors = self.errors
        return inp

    def _parse_strategies(self, ics: dict) -> list[StrategyInput]:
        strategies = []
        if not ics:
            self._warn_missing("interest_crediting_strategy")
            return strategies

        for json_name, details in self._strategy_records(ics):
            excel_name = STRATEGY_MAP.get(json_name, "")
            if not excel_name:
                # fuzzy match
                jl = json_name.lower()
                for k, v in STRATEGY_MAP.items():
                    if k.lower() in jl or jl in k.lower():
                        excel_name = v
                        break
            if not excel_name:
                msg = f"Strategy '{json_name}' could not be mapped to an Excel strategy name"
                self.errors.append(msg)
                log.warning("[%s] %s", self.pdf_name, msg)
                continue

            alloc_raw = details.get("Allocation") or details.get("allocation") or 0
            try:
                alloc = float(str(alloc_raw).replace("%", ""))
                alloc = alloc / 100 if alloc > 1 else alloc
            except ValueError:
                alloc = 0.0

            rate_type = _detect_rate_column(excel_name)

            # Detect rate value — check multiple keys
            rate_raw = None
            if rate_type == "cap":
                rate_raw = (
                    details.get("Cap")
                    or details.get("cap")
                    or details.get("Rate")
                    or details.get("rate")
                    or details.get("Participation_Rate")
                )
            elif rate_type == "parrate":
                rate_raw = (
                    details.get("Participation_Rate")
                    or details.get("par_rate")
                    or details.get("ParRate")
                    or details.get("Rate")
                    or details.get("rate")
                )
            elif rate_type == "spread":
                rate_raw = details.get("Spread") or details.get("spread")
            elif rate_type == "triggered":
                rate_raw = details.get("Triggered_Rate") or details.get("triggered_rate") or details.get("Rate")
            elif rate_type == "fixed":
                rate_raw = details.get("Rate") or details.get("rate") or details.get("Fixed_Rate")

            if rate_raw is None:
                rate_raw = details.get("Rate") or details.get("rate") or 0

            try:
                rate = float(str(rate_raw).replace("%", ""))
                rate = rate / 100 if rate > 1 else rate
            except (ValueError, TypeError):
                rate = 0.0

            strategies.append(StrategyInput(
                excel_name=excel_name,
                allocation=alloc,
                rate=rate,
                rate_type=rate_type,
            ))

        return strategies

# ---------------------------------------------------------------------------
# Phase 1: WorkbookPopulator
# ---------------------------------------------------------------------------

class WorkbookPopulator:
    """Writes TestCaseInputs into a copy of the template workbook via openpyxl."""

    # Mapping: named_range_key → value_attr on TestCaseInputs
    NAMED_RANGE_INPUTS = [
        ("Category",   "category"),
        ("Product",    "product"),
        ("LivBen",     "livben"),
        ("Summary",    "summary"),
        ("SSubCat",    "ssubcat"),
        ("MCIRInd",    "mcirind"),
        ("IssAge1",    "iss_age1"),
        ("IssAge2",    "iss_age2"),
        ("MatAge",     "mat_age"),
        ("State",      "state"),
        ("Election",   "election"),
        ("RMDInd",     "rmd_ind"),
        ("CLPrem",     "cl_prem"),
        ("PFED",       "pfed"),
        ("MWVRate",    "mwv_rate"),
        ("CGR",        "cgr"),
        ("GMIR",       "gmir"),
        ("GMABC",      "gmabc"),
        ("GMABR",      "gmabr"),
        ("WDRate",     "wd_rate"),
        ("LIADur",     "lia_dur"),
        ("SWInd",      "sw_ind"),
        ("SWMode",     "sw_mode"),
        ("SWAmount",   "sw_amount"),
        ("SWRate",     "sw_rate"),
        ("SWStart",    "sw_start"),
        ("SWEnd",      "sw_end"),
        ("AICOPrFee",  "aico_pr_fee"),
        ("OMR",        "omr"),
        ("OMaxR",      "omax_r"),
        ("ICP",        "icp"),
        ("GIRM1",      "girm1"),
        ("GIRM2",      "girm2"),
        ("NIRM1",      "nirm1"),
        ("NIRM2",      "nirm2"),
        ("AdvFee",     "adv_fee"),
        ("MaxIGP",     "max_igp"),
    ]

    # Additional fixed cells (not named ranges) for contract type
    CONTRACT_TYPE_CELL = ("Inputs & Summary", "C", 39)   # row 39 = Qualified/Non-Qualified

    def __init__(self, template_path: Path, inputs: TestCaseInputs, out_path: Path):
        self.template_path = template_path
        self.inputs = inputs
        self.out_path = out_path
        self.wb = None
        self.defined_names: dict[str, tuple[str, str, int]] = {}  # name → (sheet, col, row)

    def _parse_defined_names(self):
        """Build a lookup of defined name → (sheetname, col_letter, row) for single-cell names."""
        for dn in self.wb.defined_names:
            defn = self.wb.defined_names[dn]
            try:
                dests = list(defn.destinations)
            except Exception:
                continue
            if len(dests) == 1:
                sheet_title, coord = dests[0]
                coord_clean = coord.replace("$", "")
                # only handle single-cell references
                m = re.match(r'^([A-Z]+)(\d+)$', coord_clean)
                if m:
                    self.defined_names[dn] = (sheet_title, m.group(1), int(m.group(2)))

    def _write_named_range(self, name: str, value):
        if name not in self.defined_names:
            log.debug("Named range '%s' not found in workbook — skipping", name)
            return
        sheet_title, col, row = self.defined_names[name]
        ws = self.wb[sheet_title]
        ws[f"{col}{row}"] = value

    def _set_calc_manual(self):
        """Set workbook calculation to manual so inputs don't trigger mid-write recalc."""
        self.wb.calculation.calcMode = "manual"

    def _available_strategy_names(self) -> list[str]:
        """
        Read the product's available strategy order from STRTG.

        This avoids relying on dynamic-array spill values that openpyxl cannot
        reliably surface before Excel recalculation.
        """
        ws = self.wb["STRTG"]
        header_row = 13
        first_col = column_index_from_string("AG")
        last_col = column_index_from_string("CB")
        product_col = None

        for col_idx in range(first_col, last_col + 1):
            cell_val = ws.cell(row=header_row, column=col_idx).value
            if str(cell_val).strip() == self.inputs.product:
                product_col = col_idx
                break

        if product_col is None:
            log.warning(
                "[%s] Could not locate product '%s' in STRTG strategy table",
                self.inputs.pdf_name,
                self.inputs.product,
            )
            return []

        names: list[str] = []
        for row in range(14, 44):
            raw = ws.cell(row=row, column=product_col).value
            if raw in (None, ""):
                continue
            names.append(str(raw).strip())
        return names

    def _write_strategy_rates(self):
        """Write strategy allocations and rate values into the strategy input areas."""
        ws = self.wb["Inputs & Summary"]
        strat_map = {s.excel_name: s for s in self.inputs.strategies}
        available = self._available_strategy_names()

        # Reset the visible strategy selection area.
        for row in range(108, 138):
            ws[f"B{row}"] = None
            ws[f"{ALLOC_COL}{row}"] = 0

        # Reset the strategy rate table.
        for row in range(142, 172):
            ws[f"B{row}"] = None
            ws[f"{ALLOC_COL}{row}"] = 0
            for col in ("E", "F", "G", "H", "I"):
                ws[f"{col}{row}"] = None

        if not available:
            return

        selection_rows = {}
        for idx, name in enumerate(available):
            row = 108 + idx
            if row > 137:
                break
            ws[f"B{row}"] = name
            selection_rows[name] = row

        rate_rows = {}
        for idx, name in enumerate(available):
            row = 142 + idx
            if row > 171:
                break
            ws[f"B{row}"] = name
            rate_rows[name] = row

        for strat_name, s in strat_map.items():
            selection_row = selection_rows.get(strat_name)
            if selection_row is not None:
                ws[f"{ALLOC_COL}{selection_row}"] = s.allocation
            else:
                log.warning(
                    "[%s] Strategy '%s' was not found in selection rows for product '%s'",
                    self.inputs.pdf_name,
                    strat_name,
                    self.inputs.product,
                )

            rate_row = rate_rows.get(strat_name)
            if rate_row is None:
                log.warning(
                    "[%s] Strategy '%s' was not found in rate rows for product '%s'",
                    self.inputs.pdf_name,
                    strat_name,
                    self.inputs.product,
                )
                continue

            ws[f"{ALLOC_COL}{rate_row}"] = s.allocation
            rate_col = RATE_COL_ACTUAL.get(s.rate_type, "G")
            ws[f"{rate_col}{rate_row}"] = s.rate

    def populate(self):
        log.info("[%s] Populating workbook …", self.inputs.pdf_name)
        shutil.copy2(self.template_path, self.out_path)
        self.wb = load_workbook_compat(self.out_path)
        self._set_calc_manual()
        self._parse_defined_names()

        for nr_name, attr in self.NAMED_RANGE_INPUTS:
            val = getattr(self.inputs, attr, None)
            if val is None:
                continue
            self._write_named_range(nr_name, val)

        # Contract type is in C39 (named range not confirmed — write directly)
        ws_inp = self.wb["Inputs & Summary"]
        ws_inp["C39"] = self.inputs.contract_type

        self._write_strategy_rates()

        # Restore Summary to Specific before saving
        self._write_named_range("Summary", "Specific")

        self.wb.save(self.out_path)
        log.info("[%s] Saved → %s", self.inputs.pdf_name, self.out_path)
        self.wb.close()

# ---------------------------------------------------------------------------
# Phase 2: ReportReader
# ---------------------------------------------------------------------------

class ReportReader:
    """Reads computed output columns from the Report tab."""

    # Fixed columns in the Report tab (row 11 headers, data starts row 13)
    OUTPUT_COLS = {
        "S":  "index_change_col_S",
        "T":  "index_change_col_T",
        "U":  "index_change_col_U",
        "V":  "index_change_col_V",
        "W":  "index_change_col_W",
        "X":  "Credited_Interest_Rate",
        "Y":  "Interest_Earned",
        "AF": "Contract_Anniversary_Value",
        "AG": "Cash_Surrender_Value",
        "AH": "Minimum_Withdrawal_Value",
        "AI": "GMAB_Value",
        "R":  "Age",
        "Q":  "Year",
    }
    HEADER_ROW = 11
    DATA_START_ROW = 13

    def __init__(self, wb_path: Path):
        self.wb_path = wb_path

    def _col_index(self, col_letter: str) -> int:
        return column_index_from_string(col_letter)

    def read(self, scenario: str) -> dict:
        """Return dict of column_name → list of values for the given scenario."""
        wb = load_workbook_compat(self.wb_path, data_only=True, read_only=False)
        ws = wb["Report"]

        # Discover actual header labels from row 11 to get dynamic index change columns
        headers: dict[str, str] = {}    # col_letter → header text
        for cell in ws[self.HEADER_ROW]:
            if cell.value and isinstance(cell.value, str):
                col_ltr = get_column_letter(cell.column)
                headers[col_ltr] = cell.value.strip()

        # Build column list: use headers for index change cols (non-empty cols between R and AF)
        out: dict[str, list] = {}
        fixed_read_cols = set(self.OUTPUT_COLS.keys())

        # Dynamic index change cols (cols S..W roughly)
        dyn_index_cols = {}
        for col_ltr, hdr in headers.items():
            ci = self._col_index(col_ltr)
            if 19 <= ci <= 23:  # S=19..W=23
                if hdr and "change" in hdr.lower():
                    safe_key = re.sub(r'[^a-zA-Z0-9_]', '_', hdr)
                    dyn_index_cols[col_ltr] = safe_key

        read_cols = {**{k: v for k, v in self.OUTPUT_COLS.items()},
                     **dyn_index_cols}

        for key in read_cols.values():
            out[key] = []

        # Check for Excel errors in any cell
        errors_found = []

        saw_data = False
        for row in ws.iter_rows(min_row=self.DATA_START_ROW, values_only=False):
            row_dict = {}
            for idx, c in enumerate(row, start=1):
                row_dict[get_column_letter(idx)] = c
            year_cell = row_dict.get("Q")
            if year_cell is None or year_cell.value is None:
                if saw_data:
                    break
                continue
            saw_data = True
            # Check for formula errors
            for cl, key in read_cols.items():
                cell = row_dict.get(cl)
                if cell and isinstance(cell.value, str) and cell.value.startswith("#"):
                    errors_found.append(f"{cl}{cell.row}={cell.value}")
                val = cell.value if cell else None
                out[key].append(val)

        wb.close()

        if errors_found:
            log.error("[Report tab errors in %s scenario=%s] %s",
                      self.wb_path.name, scenario, ", ".join(errors_found[:5]))

        out["_errors"] = errors_found
        return out


SCENARIO_OUTPUT_SUFFIX = ".scenario_output.json"


def scenario_output_path_for_workbook(wb_path: Path) -> Path:
    return wb_path.with_name(f"{wb_path.stem}{SCENARIO_OUTPUT_SUFFIX}")

# ---------------------------------------------------------------------------
# Phase 2: OutputGatherer
# ---------------------------------------------------------------------------

class OutputGatherer:
    """Iterates recalculated workbooks and assembles tool_calc_output.json."""

    SCENARIOS = ["specific", "zero_growth", "constant_growth"]
    SUMMARY_MAP = {
        "specific":        "Specific",
        "zero_growth":     "Zero Growth",
        "constant_growth": "Constant Growth",
    }

    def __init__(self, output_dir: Path, error_log: list[dict]):
        self.output_dir = output_dir
        self.error_log = error_log

    def gather(self, test_cases: dict) -> dict:
        result = {}
        for pdf_name in test_cases:
            wb_path = self.output_dir / f"{pdf_name}.xlsx"
            if not wb_path.exists():
                log.warning("Workbook not found for %s — skipping output gather", pdf_name)
                continue
            result[pdf_name] = {"scenario": {}}
            scenario_output_path = scenario_output_path_for_workbook(wb_path)
            if scenario_output_path.exists():
                log.info("[%s] Loading captured scenario outputs from %s", pdf_name, scenario_output_path.name)
                captured = load_json(scenario_output_path)
                scenario_map = captured.get("scenario", {})
                for scenario in self.SCENARIOS:
                    log.info("[%s] Reading captured Report output for scenario=%s", pdf_name, scenario)
                    data = copy.deepcopy(scenario_map.get(scenario, {}))
                    errs = data.pop("_errors", [])
                    if errs:
                        msg = (f"report tab has error in illustration tool, please check calculation "
                               f"({', '.join(errs[:3])})")
                        self.error_log.append({"test_case": pdf_name, "error_message": msg})
                    result[pdf_name]["scenario"][scenario] = data
                continue

            reader = ReportReader(wb_path)
            for scenario in self.SCENARIOS:
                log.info("[%s] Reading Report tab for scenario=%s", pdf_name, scenario)
                data = reader.read(scenario)
                errs = data.pop("_errors", [])
                if errs:
                    msg = (f"report tab has error in illustration tool, please check calculation "
                           f"({', '.join(errs[:3])})")
                    self.error_log.append({"test_case": pdf_name, "error_message": msg})
                result[pdf_name]["scenario"][scenario] = data
        return result

# ---------------------------------------------------------------------------
# Phase 3: OutputComparator
# ---------------------------------------------------------------------------

class OutputComparator:
    """Compares tool_calc_output.json vs product_structure.json."""

    SCENARIO_JSON_MAP = {
        "specific":        "specific",
        "zero_growth":     "zero_growth",
        "constant_growth": "constant_growth",
    }
    TOOL_TO_REF_COLUMN_MAP = {
        "Age":                        "Age",
        "Credited_Interest_Rate":     "Credited_Interest_Rate",
        "Interest_Earned":            "Interest_Earned",
        "Contract_Anniversary_Value": "Contract_Anniversary_Value",
        "Cash_Surrender_Value":       "Cash_Surrender_Value",
        "Minimum_Withdrawal_Value":   "Death_Benefit",
        "GMAB_Value":                 "Minimum_Accumulation_Value",
    }
    RATE_ABS_TOLERANCE = 0.0005
    VALUE_ABS_TOLERANCE = 1.0

    def __init__(self, tool_output: dict, product_structure: dict):
        self.tool = tool_output
        self.ref = product_structure
        self.records: list[dict] = []
        self._counter = 1

    def _log(self, pdf_name: str, scenario: str, message: str):
        self.records.append({
            "number":    self._counter,
            "test_case": pdf_name,
            "scenario":  scenario,
            "message":   message,
        })
        self._counter += 1

    @staticmethod
    def _normalized_key(value: str) -> str:
        return re.sub(r"[^a-z0-9]+", "", str(value).lower())

    def _find_ref_column(self, tool_col: str, ref_scenario: dict) -> str | None:
        preferred = self.TOOL_TO_REF_COLUMN_MAP.get(tool_col)
        if preferred in ref_scenario:
            return preferred

        normalized_tool = self._normalized_key(tool_col)
        normalized_preferred = self._normalized_key(preferred or "")

        for ref_col in ref_scenario:
            normalized_ref = self._normalized_key(ref_col)
            if normalized_ref == normalized_tool:
                return ref_col
            if normalized_preferred and normalized_ref == normalized_preferred:
                return ref_col

        for ref_col in ref_scenario:
            normalized_ref = self._normalized_key(ref_col)
            if normalized_ref in normalized_tool or normalized_tool in normalized_ref:
                return ref_col
            if normalized_preferred and (
                normalized_ref in normalized_preferred or normalized_preferred in normalized_ref
            ):
                return ref_col
        return None

    @staticmethod
    def _normalize_value(value: Any) -> tuple[str, Any]:
        if value is None:
            return ("empty", None)
        if isinstance(value, bool):
            return ("text", str(value))
        if isinstance(value, (int, float)):
            return ("number", float(value))

        text = str(value).strip()
        if not text:
            return ("empty", None)
        if text in {"-", "--"}:
            return ("dash", text)
        if text.upper() == "N/A":
            return ("text", "N/A")

        cleaned = text.replace(",", "").replace("$", "")
        negative = cleaned.startswith("(") and cleaned.endswith(")")
        if negative:
            cleaned = cleaned[1:-1]
        is_percent = cleaned.endswith("%")
        if is_percent:
            cleaned = cleaned[:-1]

        try:
            num = float(cleaned)
            if negative:
                num = -num
            if is_percent:
                num /= 100
            return ("number", num)
        except ValueError:
            return ("text", text)

    def _values_match(self, column_name: str, tool_value: Any, ref_value: Any) -> bool:
        tool_kind, tool_normalized = self._normalize_value(tool_value)
        ref_kind, ref_normalized = self._normalize_value(ref_value)

        if tool_kind == ref_kind == "empty":
            return True
        if tool_kind == ref_kind == "number":
            normalized_col = self._normalized_key(column_name)
            if normalized_col in {"year", "age"}:
                return int(round(tool_normalized)) == int(round(ref_normalized))
            tolerance = (
                self.RATE_ABS_TOLERANCE
                if any(token in normalized_col for token in ("rate", "change", "percentage", "credit"))
                else self.VALUE_ABS_TOLERANCE
            )
            return abs(tool_normalized - ref_normalized) <= tolerance
        return tool_normalized == ref_normalized

    @staticmethod
    def _row_label(years: list[Any], ages: list[Any], idx: int) -> str:
        year = years[idx] if idx < len(years) else None
        age = ages[idx] if idx < len(ages) else None
        parts = []
        if year not in (None, ""):
            parts.append(f"Year={year}")
        if age not in (None, ""):
            parts.append(f"Age={age}")
        return ", ".join(parts) if parts else f"row {idx + 1}"

    def _should_skip_column(self, column_name: str) -> bool:
        normalized = self._normalized_key(column_name)
        return normalized == "year" or normalized.startswith("indexchangecol")

    def compare(self):
        for pdf_name, tool_data in self.tool.items():
            ref_tc = self.ref.get(pdf_name, {})
            ref_scenarios = ref_tc.get("scenario", ref_tc)

            for scenario, scen_data in tool_data.get("scenario", {}).items():
                ref_scenario = ref_scenarios.get(scenario, {})
                if not ref_scenario:
                    self._log(pdf_name, scenario, "scenario is missing in product_structure.json")
                    continue

                tool_years = scen_data.get("Year", [])
                ref_years = ref_scenario.get("Year", [])
                tool_ages = scen_data.get("Age", [])
                ref_ages = ref_scenario.get("Age", ref_scenario.get("age", []))

                for tool_col, tool_vals in scen_data.items():
                    if tool_col == "_errors" or not isinstance(tool_vals, list):
                        continue
                    if self._should_skip_column(tool_col):
                        continue

                    ref_col = self._find_ref_column(tool_col, ref_scenario)
                    if ref_col is None:
                        self._log(pdf_name, scenario, f"column {tool_col} is missing in product_structure.json")
                        continue

                    ref_vals = ref_scenario.get(ref_col)
                    if not isinstance(ref_vals, list):
                        self._log(
                            pdf_name,
                            scenario,
                            f"column {ref_col} in product_structure.json is not a comparable list",
                        )
                        continue

                    if len(tool_vals) != len(ref_vals):
                        self._log(
                            pdf_name,
                            scenario,
                            f"column {tool_col} length mismatch (tool={len(tool_vals)}, ref={len(ref_vals)})",
                        )

                    shared_len = min(len(tool_vals), len(ref_vals))
                    for idx in range(shared_len):
                        tool_kind, tool_normalized = self._normalize_value(tool_vals[idx])
                        ref_kind, _ = self._normalize_value(ref_vals[idx])
                        normalized_col = self._normalized_key(tool_col)
                        if (
                            idx == 0
                            and tool_kind == "number"
                            and tool_normalized == 0
                            and ref_kind in {"dash", "empty"}
                            and any(token in normalized_col for token in ("rate", "change", "interest"))
                        ):
                            continue
                        if self._values_match(tool_col, tool_vals[idx], ref_vals[idx]):
                            continue
                        row_label = self._row_label(
                            tool_years if tool_years else ref_years,
                            tool_ages if tool_ages else ref_ages,
                            idx,
                        )
                        self._log(
                            pdf_name,
                            scenario,
                            f"column {tool_col} mismatch at {row_label} (tool={tool_vals[idx]!r}, ref={ref_vals[idx]!r})",
                        )
                        break
        return self.records

# ---------------------------------------------------------------------------
# Recalc helper writer
# ---------------------------------------------------------------------------

RECALC_HELPER = '''"""
recalc_helper.py  —  uses xlwings to open, switch scenario, recalculate and save
each workbook in the output folder.

Usage:
    python recalc_helper.py ./excel_test_cases
    python recalc_helper.py ./excel_test_cases "example.pdf.xlsx" "another.pdf.xlsx"
"""
import json
import re
import shutil
import sys
import tempfile
import time
import uuid
from pathlib import Path

try:
    import xlwings as xw
except ImportError:
    print("xlwings not installed.  Run: pip install xlwings")
    sys.exit(1)

try:
    sys.stdout.reconfigure(line_buffering=True)
except Exception:
    pass

SCENARIO_NAMED_RANGE = "Summary"
SCENARIOS = {
    "specific":        "Specific",
    "zero_growth":     "Zero Growth",
    "constant_growth": "Constant Growth",
}
OUTPUT_COLS = {
    "S":  "index_change_col_S",
    "T":  "index_change_col_T",
    "U":  "index_change_col_U",
    "V":  "index_change_col_V",
    "W":  "index_change_col_W",
    "X":  "Credited_Interest_Rate",
    "Y":  "Interest_Earned",
    "AF": "Contract_Anniversary_Value",
    "AG": "Cash_Surrender_Value",
    "AH": "Minimum_Withdrawal_Value",
    "AI": "GMAB_Value",
    "R":  "Age",
    "Q":  "Year",
}
HEADER_ROW = 11
DATA_START_ROW = 13

def _app_visible_default() -> bool:
    # Excel startup on Windows is often more reliable when the app is visible.
    return sys.platform.startswith("win")

def _configure_app(app):
    if sys.platform.startswith("win"):
        print("[recalc] Skipping advanced Excel startup configuration on Windows.", flush=True)
        return

    # Reduce UI prompts during batch automation.
    try:
        app.display_alerts = False
    except Exception:
        pass
    try:
        app.screen_updating = False
    except Exception:
        pass
    try:
        app.calculation = "automatic"
    except Exception:
        pass
    try:
        app.api.enable_events = False
    except Exception:
        pass
    try:
        app.api.ask_to_update_links = False
    except Exception:
        pass

def _resolve_requested_workbooks(folder: Path, requested: list[str]) -> list[Path]:
    if not requested:
        return sorted(folder.glob("*.xlsx"))

    resolved = []
    missing = []
    for name in requested:
        raw_name = Path(name).name
        candidates = []
        if raw_name.lower().endswith(".xlsx"):
            candidates.append(folder / raw_name)
        else:
            candidates.append(folder / f"{raw_name}.xlsx")
            candidates.append(folder / raw_name)

        workbook_path = next((candidate for candidate in candidates if candidate.exists()), None)
        if workbook_path is None:
            missing.append(raw_name)
            continue
        resolved.append(workbook_path)

    if missing:
        print("Requested workbook(s) not found:", ", ".join(missing))
        sys.exit(1)

    return resolved

def _excel_staging_root() -> Path:
    if sys.platform.startswith("win"):
        root = Path(tempfile.gettempdir()) / "product_illustration_automation_recalc"
        root.mkdir(parents=True, exist_ok=True)
        return root

    candidates = [
        Path.home() / "Library/Containers/com.microsoft.Excel/Data/Documents",
        Path.home() / "Library/Group Containers/UBF8T346G9.Office",
    ]
    if sys.platform == "darwin":
        for candidate in candidates:
            if candidate.exists():
                root = candidate / "product_illustration_automation_recalc"
                root.mkdir(parents=True, exist_ok=True)
                return root
        raise RuntimeError(
            "Could not locate an Excel sandbox-safe staging folder. "
            "Expected one of: ~/Library/Containers/com.microsoft.Excel/Data/Documents "
            "or ~/Library/Group Containers/UBF8T346G9.Office"
        )

    root = Path(tempfile.gettempdir()) / "product_illustration_automation_recalc"
    root.mkdir(parents=True, exist_ok=True)
    return root

def _prepare_staged_copy(original_path: Path, staging_dir: Path) -> Path:
    staged_path = staging_dir / original_path.name
    shutil.copy2(original_path, staged_path)
    return staged_path

def _col_index(col_letter: str) -> int:
    idx = 0
    for char in col_letter.upper():
        idx = idx * 26 + (ord(char) - ord("A") + 1)
    return idx

def _scenario_output_path(xl_path: Path) -> Path:
    return xl_path.with_name(f"{xl_path.stem}.scenario_output.json")

def _capture_report_data(wb, scenario: str) -> dict:
    ws = wb.sheets["Report"]
    last_col = int(ws.used_range.last_cell.column)
    last_row = int(ws.used_range.last_cell.row)

    header_values = ws.range((HEADER_ROW, 1), (HEADER_ROW, last_col)).value
    if not isinstance(header_values, list):
        header_values = [header_values]

    headers = {}
    for col_idx, value in enumerate(header_values, start=1):
        if isinstance(value, str) and value.strip():
            headers[col_idx] = value.strip()

    read_cols = {_col_index(col_letter): key for col_letter, key in OUTPUT_COLS.items()}
    for col_idx, header in headers.items():
        if 19 <= col_idx <= 23 and "change" in header.lower():
            safe_key = re.sub(r"[^a-zA-Z0-9_]", "_", header)
            read_cols[col_idx] = safe_key

    out = {key: [] for key in read_cols.values()}
    errors_found = []
    year_col_idx = _col_index("Q")

    saw_data = False
    for row_idx in range(DATA_START_ROW, last_row + 1):
        year_val = ws.range((row_idx, year_col_idx)).value
        if year_val is None:
            if saw_data:
                break
            continue
        saw_data = True
        for col_idx, key in read_cols.items():
            val = ws.range((row_idx, col_idx)).value
            if isinstance(val, str) and val.startswith("#"):
                errors_found.append(f"{col_idx}:{row_idx}={val}")
            out[key].append(val)

    if errors_found:
        print(
            f"[recalc] Report tab errors in {wb.name} scenario={scenario}: {', '.join(errors_found[:5])}",
            flush=True,
        )

    out["_errors"] = errors_found
    return out

def recalc_workbook(app, xl_path: Path, staging_dir: Path):
    print(f"[recalc] Opening {xl_path.name} ...", flush=True)
    wb = None
    staged_path = _prepare_staged_copy(xl_path, staging_dir)
    scenario_output = {"scenario": {}}
    try:
        # update_links=False avoids repeated "update links/grant access" prompts
        # when files contain references.
        try:
            wb = app.books.open(
                str(staged_path.resolve()),
                update_links=False,
                read_only=False,
                notify=False,
                add_to_mru=False,
            )
        except TypeError:
            # Older xlwings builds may not support all kwargs.
            wb = app.books.open(str(staged_path.resolve()))

        for scenario_key, excel_val in SCENARIOS.items():
            print(f"  scenario={scenario_key} ({excel_val})", flush=True)
            # Write Summary named range
            rng = wb.names[SCENARIO_NAMED_RANGE].refers_to_range
            rng.value = excel_val
            # Force recalculate
            wb.app.calculate()
            time.sleep(3)   # give Excel time to finish
            scenario_output["scenario"][scenario_key] = _capture_report_data(wb, scenario_key)

        # Leave as Specific before final save
        rng = wb.names[SCENARIO_NAMED_RANGE].refers_to_range
        rng.value = "Specific"
        wb.app.calculate()
        time.sleep(2)
        wb.save()
        shutil.copy2(staged_path, xl_path)
        scenario_output_path = _scenario_output_path(xl_path)
        scenario_output_path.write_text(
            json.dumps(scenario_output, indent=2, default=str),
            encoding="utf-8",
        )
        print(f"  Captured scenario outputs -> {scenario_output_path.name}", flush=True)
        print(f"  Saved {xl_path.name}", flush=True)
    finally:
        if wb is not None:
            wb.close()
        try:
            staged_path.unlink()
        except FileNotFoundError:
            pass

if __name__ == "__main__":
    folder = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(".")
    requested_workbooks = sys.argv[2:]
    xlsx_files = _resolve_requested_workbooks(folder, requested_workbooks)
    if not xlsx_files:
        print(f"No .xlsx files found in {folder}", flush=True)
        sys.exit(1)

    visible = _app_visible_default()
    print(f"[recalc] Starting Excel app ... visible={visible}", flush=True)
    app = xw.App(visible=visible, add_book=False)
    print("[recalc] Excel app started.", flush=True)
    _configure_app(app)
    print("[recalc] Excel app ready.", flush=True)
    staging_dir = _excel_staging_root() / f"run_{folder.name}_{uuid.uuid4().hex[:8]}"
    staging_dir.mkdir(parents=True, exist_ok=True)
    try:
        for f in xlsx_files:
            recalc_workbook(app, f, staging_dir)
    finally:
        app.quit()
        shutil.rmtree(staging_dir, ignore_errors=True)
    print("Done.", flush=True)
'''

# ---------------------------------------------------------------------------
# Main pipeline
# ---------------------------------------------------------------------------

def load_json(path: Path) -> dict:
    with open(path, encoding="utf-8") as fh:
        return json.load(fh)

def save_json(data: dict, path: Path):
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(data, fh, indent=2, default=str)
    log.info("Saved %s", path)

def save_csv(records: list[dict], path: Path, fieldnames: list[str] | None = None):
    if not records and not fieldnames:
        log.info("No records to write for %s", path)
        return
    with open(path, "w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames or list(records[0].keys()))
        writer.writeheader()
        if records:
            writer.writerows(records)
    log.info("Saved %s (%d rows)", path, len(records))

def _resolve_input_path(raw_value: str | Path | None, default_path: Path) -> Path:
    if raw_value in (None, ""):
        return default_path

    path = Path(raw_value).expanduser()
    if path.is_absolute():
        return path

    candidates = [
        Path.cwd() / path,
        SCRIPT_DIR / path,
        PROJECT_PARENT_DIR / path,
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return candidates[0]


def _resolve_optional_input_path(raw_value: str | Path | None) -> Path | None:
    if raw_value in (None, ""):
        return None

    path = Path(raw_value).expanduser()
    if path.is_absolute():
        return path

    candidates = [
        Path.cwd() / path,
        SCRIPT_DIR / path,
        PROJECT_PARENT_DIR / path,
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return candidates[0]

def _resolve_output_path(raw_value: str | Path | None) -> Path | None:
    if raw_value in (None, ""):
        return None

    path = Path(raw_value).expanduser()
    if path.is_absolute():
        return path
    return (Path.cwd() / path).resolve()

def _normalize_test_case_values(raw_value: Any) -> list[str]:
    if raw_value in (None, "", []):
        return []

    if isinstance(raw_value, str):
        parts = [part.strip() for part in raw_value.split(",")]
        return [part for part in parts if part]

    if isinstance(raw_value, (list, tuple, set)):
        normalized = []
        for item in raw_value:
            if item is None:
                continue
            text = str(item).strip()
            if text:
                normalized.append(text)
        return normalized

    text = str(raw_value).strip()
    return [text] if text else []

def _test_case_aliases(name: str) -> set[str]:
    path_name = Path(name).name
    aliases = {
        name.casefold(),
        path_name.casefold(),
        Path(name).stem.casefold(),
        Path(path_name).stem.casefold(),
    }
    return {alias for alias in aliases if alias}

def filter_product_structure(product_structure: dict, requested_cases: list[str]) -> dict:
    if not requested_cases:
        return product_structure

    alias_to_actual: dict[str, str] = {}
    for actual_name in product_structure:
        for alias in _test_case_aliases(actual_name):
            alias_to_actual.setdefault(alias, actual_name)

    filtered: dict[str, Any] = {}
    missing: list[str] = []

    for requested in requested_cases:
        actual_name = None
        for alias in _test_case_aliases(requested):
            actual_name = alias_to_actual.get(alias)
            if actual_name:
                break
        if not actual_name:
            missing.append(requested)
            continue
        filtered[actual_name] = product_structure[actual_name]

    if missing:
        raise ValueError(
            "Requested test case(s) not found in product_structure.json: "
            + ", ".join(missing)
        )

    return filtered

def normalize_phases(raw_phases: Any) -> list[int]:
    phases = raw_phases if raw_phases not in (None, []) else DEFAULT_PHASES
    normalized = [int(phase) for phase in phases]
    invalid = [phase for phase in normalized if phase not in VALID_PHASES]
    if invalid:
        raise ValueError(f"Unsupported phase(s): {invalid}. Valid values are 1, 2, 3.")
    return normalized

def create_run_output_dir(results_root: Path) -> tuple[Path, str]:
    run_date = dt.datetime.now().astimezone().strftime("%Y-%m-%d")
    valuation_id = f"valuation_{dt.datetime.now().astimezone().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}"
    output_dir = results_root / run_date / valuation_id
    output_dir.mkdir(parents=True, exist_ok=False)
    return output_dir, valuation_id

def write_run_manifest(
    output_dir: Path,
    workbook_dir: Path,
    valuation_id: str,
    template_path: Path,
    json_path: Path,
    phases: list[int],
    requested_cases: list[str],
    pdf_dir: Path | None = None,
    pdf_extractor: str | None = None,
    log_path: Path | None = None,
):
    manifest = {
        "submitted_at": dt.datetime.now().astimezone().isoformat(),
        "valuation_id": valuation_id,
        "template_path": str(template_path),
        "json_path": str(json_path),
        "pdf_dir": str(pdf_dir) if pdf_dir else None,
        "pdf_extractor": pdf_extractor,
        "phases": phases,
        "test_cases": requested_cases,
        "output_dir": str(output_dir),
        "workbook_dir": str(workbook_dir),
        "run_log_path": str(log_path) if log_path else None,
    }
    save_json(manifest, output_dir / "run_config.json")

def run_recalc_helper(recalc_path: Path, workbook_dir: Path, workbook_names: list[str]):
    cmd = [sys.executable, "-u", str(recalc_path), str(workbook_dir)]
    cmd.extend(workbook_names)
    log.info("===== RECALCULATION: Recalculate workbooks =====")
    log.info("Running: %s", " ".join(cmd))
    subprocess.run(cmd, check=True)

def run_phase1(template_path: Path, product_structure: dict, output_dir: Path) -> list[dict]:
    output_dir.mkdir(parents=True, exist_ok=True)
    error_log: list[dict] = []
    counter = 1

    for pdf_name, tc in product_structure.items():
        log.info("=== Phase 1: %s ===", pdf_name)
        parser = TestCaseParser(pdf_name, tc)
        inputs = parser.parse()

        for err in inputs.errors:
            error_log.append({"number": counter, "test_case": pdf_name, "error_message": err})

        out_path = output_dir / f"{pdf_name}.xlsx"
        populator = WorkbookPopulator(template_path, inputs, out_path)
        try:
            populator.populate()
        except Exception as exc:
            msg = f"Failed to populate workbook: {exc}"
            log.error("[%s] %s", pdf_name, msg)
            error_log.append({"number": counter, "test_case": pdf_name, "error_message": msg})

        counter += 1

    return error_log

def run_phase2(output_dir: Path, product_structure: dict) -> tuple[dict, list[dict]]:
    error_log: list[dict] = []
    gatherer = OutputGatherer(output_dir, error_log)
    tool_output = gatherer.gather(product_structure)
    return tool_output, error_log

def run_phase3(tool_output: dict, product_structure: dict) -> list[dict]:
    comparator = OutputComparator(tool_output, product_structure)
    return comparator.compare()

def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    ap = argparse.ArgumentParser(description="Annuity Automation Script")
    ap.add_argument("--template", help="Path to template .xlsx")
    ap.add_argument("--json", help="Path to product_structure.json")
    ap.add_argument("--pdf-dir", help="Directory containing source PDFs for integrated extraction + automation flow")
    ap.add_argument(
        "--pdf-extractor",
        choices=["auto", "pypdf", "pypdf2", "pdfkit"],
        help="PDF text extraction backend when --pdf-dir is used",
    )
    ap.add_argument("--output-dir",
                    help="Folder for this run's output files. If omitted, a dated valuation folder is created under results/.")
    ap.add_argument("--results-root",
                    help="Base results folder for auto-created dated valuation folders "
                         f"(default: {DEFAULT_RESULTS_ROOT})")
    ap.add_argument("--phases", nargs="+", type=int,
                    help="Which phases to run: 1, 2, 3 (default: all)")
    ap.add_argument("--test-cases", nargs="+",
                    help="Run only the specified PDF test case names. Accepts one or more file names.")
    return ap.parse_args(argv)

def build_runtime_settings(args: argparse.Namespace, runtime_config: dict[str, Any] | None) -> dict[str, Any]:
    runtime_config = runtime_config or {}
    cli_values = {
        "template": args.template,
        "json": args.json,
        "pdf_dir": args.pdf_dir,
        "pdf_extractor": args.pdf_extractor,
        "output_dir": args.output_dir,
        "results_root": args.results_root,
        "phases": args.phases,
        "test_cases": args.test_cases,
    }

    merged: dict[str, Any] = {}
    merged.update(runtime_config)
    for key, value in cli_values.items():
        if value is not None:
            merged[key] = value

    template_path = _resolve_input_path(merged.get("template"), DEFAULT_TEMPLATE_PATH)
    json_path = _resolve_input_path(merged.get("json"), DEFAULT_JSON_PATH)
    pdf_dir = _resolve_optional_input_path(merged.get("pdf_dir"))
    pdf_extractor = str(merged.get("pdf_extractor") or "auto")
    explicit_output_dir = _resolve_output_path(merged.get("output_dir"))
    results_root = _resolve_output_path(merged.get("results_root")) or DEFAULT_RESULTS_ROOT
    phases = normalize_phases(merged.get("phases"))
    requested_cases = _normalize_test_case_values(merged.get("test_cases"))

    if explicit_output_dir is None and 1 not in phases and any(phase in phases for phase in (2, 3)):
        raise ValueError(
            "When running phase 2 and/or 3 without phase 1, please provide --output-dir "
            "or config['output_dir'] for an existing run folder."
        )

    if explicit_output_dir is None:
        output_dir, valuation_id = create_run_output_dir(results_root)
    else:
        output_dir = explicit_output_dir
        output_dir.mkdir(parents=True, exist_ok=True)
        valuation_id = output_dir.name

    workbook_dir = output_dir / "excel_test_cases"
    if explicit_output_dir is not None and not workbook_dir.exists():
        existing_root_workbooks = list(output_dir.glob("*.xlsx"))
        if existing_root_workbooks:
            workbook_dir = output_dir

    return {
        "template_path": template_path,
        "json_path": json_path,
        "pdf_dir": pdf_dir,
        "pdf_extractor": pdf_extractor,
        "output_dir": output_dir,
        "workbook_dir": workbook_dir,
        "results_root": results_root,
        "phases": phases,
        "test_cases": requested_cases,
        "valuation_id": valuation_id,
    }

def main(runtime_config: dict[str, Any] | None = None):
    configure_logging()
    started_at = time.perf_counter()
    args = parse_args()

    try:
        settings = build_runtime_settings(args, runtime_config if runtime_config is not None else config)
    except ValueError as exc:
        log.error("%s", exc)
        sys.exit(1)

    template_path = settings["template_path"]
    json_path = settings["json_path"]
    pdf_dir = settings["pdf_dir"]
    pdf_extractor = settings["pdf_extractor"]
    output_dir = settings["output_dir"]
    workbook_dir = settings["workbook_dir"]
    phases = settings["phases"]
    requested_cases = settings["test_cases"]
    valuation_id = settings["valuation_id"]
    run_log_path = output_dir / "run.log"

    configure_logging(run_log_path)

    if 1 in phases and not template_path.exists():
        log.error("Template file not found: %s", template_path)
        sys.exit(1)
    if pdf_dir is None and not json_path.exists():
        log.error("JSON file not found: %s", json_path)
        sys.exit(1)
    if pdf_dir is not None and not pdf_dir.exists():
        log.error("PDF directory not found: %s", pdf_dir)
        sys.exit(1)

    log.info("Run output directory: %s", output_dir)
    log.info("Workbook directory: %s", workbook_dir)
    log.info("Valuation ID: %s", valuation_id)
    log.info("Run log: %s", run_log_path)

    if pdf_dir is not None:
        json_path = output_dir / "product_structure.json"
        drop_file = output_dir / "drop_pdf"
        log.info("===== EXTRACTION: Parse PDFs into product_structure.json =====")
        try:
            extractor = AnnuityExtractorPipeline(
                pdf_dir=pdf_dir,
                output_json=json_path,
                drop_file=drop_file,
                extractor_backend=pdf_extractor,
            )
            product_structure, dropped = extractor.run()
            log.info("Extraction complete. drop_pdf count: %d", len(dropped))
        except RuntimeError as exc:
            log.error("Error initializing PDF extractor: %s", exc)
            sys.exit(2)
    else:
        product_structure = load_json(json_path)
    try:
        product_structure = filter_product_structure(product_structure, requested_cases)
    except ValueError as exc:
        log.error("%s", exc)
        sys.exit(1)

    workbook_names = [f"{pdf_name}.xlsx" for pdf_name in product_structure]
    all_errors: list[dict] = []
    tool_output: dict = {}

    if requested_cases:
        log.info("Selected test case(s): %s", ", ".join(product_structure.keys()))

    write_run_manifest(
        output_dir=output_dir,
        workbook_dir=workbook_dir,
        valuation_id=valuation_id,
        template_path=template_path,
        json_path=json_path,
        phases=phases,
        requested_cases=list(product_structure.keys()),
        pdf_dir=pdf_dir,
        pdf_extractor=pdf_extractor,
        log_path=run_log_path,
    )

    recalc_path = workbook_dir / "recalc_helper.py"
    workbook_dir.mkdir(parents=True, exist_ok=True)
    recalc_path.write_text(RECALC_HELPER, encoding="utf-8")
    log.info("recalc_helper.py written to %s", recalc_path)

    if 1 in phases:
        log.info("===== PHASE 1: Populate workbooks =====")
        errs = run_phase1(template_path, product_structure, workbook_dir)
        all_errors.extend(errs)
        for i, rec in enumerate(all_errors, 1):
            rec["number"] = i
        save_csv(all_errors, output_dir / "error_report.csv",
                 fieldnames=["number", "test_case", "error_message"])
        log.info("Phase 1 complete. %d error(s) logged.", len(all_errors))

    if 1 in phases and any(phase in phases for phase in (2, 3)):
        try:
            run_recalc_helper(recalc_path, workbook_dir, workbook_names)
        except subprocess.CalledProcessError as exc:
            log.error("Recalculation failed with exit code %s", exc.returncode)
            log.error("You can retry manually with: python %s %s", recalc_path, workbook_dir)
            sys.exit(exc.returncode or 1)

    if 2 in phases:
        log.info("===== PHASE 2: Gather outputs =====")
        tool_output, gather_errors = run_phase2(workbook_dir, product_structure)
        all_errors.extend(gather_errors)
        save_json(tool_output, output_dir / "tool_calc_output.json")

    if 3 in phases:
        log.info("===== PHASE 3: Compare outputs =====")
        if not tool_output:
            tool_calc_path = output_dir / "tool_calc_output.json"
            if tool_calc_path.exists():
                tool_output = load_json(tool_calc_path)
            else:
                log.error("tool_calc_output.json not found — run Phase 2 first")
                sys.exit(1)

        check_records = run_phase3(tool_output, product_structure)
        save_csv(check_records, output_dir / "check_report.csv",
                 fieldnames=["number", "test_case", "scenario", "message"])
        log.info("Phase 3 complete. %d check record(s).", len(check_records))

    for i, rec in enumerate(all_errors, 1):
        rec["number"] = i
    save_csv(all_errors, output_dir / "error_report.csv",
             fieldnames=["number", "test_case", "error_message"])
    log.info("All done.")
    log.info("Total processing time: %.2f seconds", time.perf_counter() - started_at)


if __name__ == "__main__":
    main()
